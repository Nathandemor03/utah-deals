import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os, shutil

# ── Load template ──────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook('/Users/nathandemordaunt/Downloads/Assignment 6 Asset Register Sheet.xlsx')
ws = wb.active

# Clear existing data rows (rows 7+)
for row in ws.iter_rows(min_row=7, max_row=ws.max_row):
    for cell in row:
        cell.value = None

# Re-write the header row with our 18 column names (keep row 6)
headers = [
    "Entry No.",
    "Description of Asset",
    "Year of Purchase",
    "Location",
    "Make and Model",
    "Serial No.",
    "Asset No.",
    "Quantity",
    "Depreciable Life (yrs)",
    "Cost of Asset ($)",
    "Year of Update",
    "Depreciation Amount ($/yr)",
    "Current Value ($)",
    "Condition of Asset (G/F/P)",
    "Expected Replacement Year",
    "2016 Replacement Cost ($)",
    "Expected Salvage ($)",
    "Comments",
]
for col_idx, hdr in enumerate(headers, 1):
    cell = ws.cell(row=6, column=col_idx)
    cell.value = hdr
    cell.font = Font(name="Arial", size=10, bold=True)
    cell.alignment = Alignment(horizontal="center", wrap_text=True)

# ── Asset data ─────────────────────────────────────────────────────────────────
# Columns in template order:
#  A  Entry No.
#  B  Description
#  C  Year of Purchase
#  D  Location
#  E  Make and Model
#  F  Serial No.
#  G  Asset No.
#  H  Quantity
#  I  Depreciable Life (yrs)
#  J  Cost of Asset ($)
#  K  Year of Update  (always 2015)
#  L  Depreciation Amount (formula)
#  M  Current Value (formula)
#  N  Condition (G/F/P)
#  O  Expected Replacement Year
#  P  2016 Replacement Cost ($)
#  Q  Expected Salvage ($)
#  R  Comments

# Each tuple: (desc, yr_purchase, location, make_model, serial, asset_no,
#              qty, dep_life, cost, condition, repl_year, repl_cost_2016, salvage, comments)
assets = [
    # ── HVAC (HM-1000s) ──────────────────────────────────────────────────────
    ("Gas Furnace", 2003, "Utility Room", "Carrier 58CVA080",
     "SN-5832ABCD", "HM-1001",
     1, 20, 2800, "F", 2018, 3200, 0,
     "Heat exchanger shows wear; recommend inspection before winter"),
    ("AC Condenser Unit", 2005, "Exterior South Side", "Trane XR13 2-Ton",
     "SN-7741TRNX", "HM-1002",
     1, 15, 2200, "F", 2020, 2600, 0,
     "Evaporator coil not inspected; condenser fins slightly bent"),

    # ── Plumbing (HM-2000s) ──────────────────────────────────────────────────
    ("Gas Water Heater", 2008, "Utility Room", "Rheem MR50245",
     "SN-4421RHWH", "HM-2001",
     1, 12, 650, "G", 2020, 750, 0,
     "No leaks; anode rod replaced 2013"),
    ("Expansion Tank", 2008, "Utility Room", "Amtrol ST-5",
     "SN-8832AMST", "HM-2002",
     1, 15, 120, "G", 2023, 150, 0,
     "Installed with water heater; pressure normal"),
    ("Water Softener", 2006, "Utility Room", "Kenmore 38300",
     "SN-3389KNSF", "HM-2003",
     1, 15, 900, "F", 2021, 1050, 0,
     "Resin tank aging; salt bridging occurred once in 2014"),

    # ── Ceiling Fans (HM-3000s) ──────────────────────────────────────────────
    ("Ceiling Fan", 2004, "Master Bedroom", "Hunter 52-in Original",
     "SN-1122HNFN", "HM-3001",
     1, 15, 180, "F", 2019, 200, 0,
     "Light kit flickers occasionally; wobble on high speed"),
    ("Ceiling Fan", 2004, "Living Room", "Hunter 52-in Original",
     "SN-1123HNFN", "HM-3002",
     1, 15, 180, "G", 2019, 200, 0,
     "Operates normally; blades recently dusted"),
    ("Ceiling Fan", 2007, "Guest Bedroom", "Hampton Bay Midili 44-in",
     "SN-5543HBMF", "HM-3003",
     1, 15, 130, "G", 2022, 150, 0,
     "No issues noted"),
    ("Ceiling Fan", 2007, "Kitchen", "Hampton Bay Midili 44-in",
     "SN-5544HBMF", "HM-3004",
     1, 15, 130, "G", 2022, 150, 0,
     "No issues noted"),

    # ── Vehicles (HM-4000s) ──────────────────────────────────────────────────
    ("Automobile", 2010, "Garage", "Toyota Camry LE 2010",
     "SN-1HGCM82633A001234", "HM-4001",
     1, 12, 21000, "G", 2022, 19000, 2000,
     "85,000 miles; routine maintenance up to date"),
    ("Bicycle", 2009, "Garage", "Trek 7.3 FX",
     "SN-WTU123456", "HM-4002",
     1, 15, 550, "G", 2024, 600, 50,
     "Tires replaced 2012; derailleur adjusted 2014"),

    # ── Kitchen Appliances (HM-5000s) ────────────────────────────────────────
    ("Gas Range / Stove", 2005, "Kitchen", "GE JGBS66REKSS",
     "SN-GE4421JGBS", "HM-5001",
     1, 15, 700, "F", 2020, 850, 0,
     "Two burner igniters need replacement; oven calibrated 2013"),
    ("Refrigerator", 2008, "Kitchen", "Whirlpool WRF560SMYB",
     "SN-WP7734RFMB", "HM-5002",
     1, 15, 1100, "G", 2023, 1250, 0,
     "Ice maker working; seals in good condition"),
    ("Microwave Oven", 2011, "Kitchen", "Panasonic NN-SN651B",
     "SN-PA9921SNMW", "HM-5003",
     1, 10, 150, "G", 2021, 175, 0,
     "Over-the-range unit; vent fan slightly noisy"),
    ("Dishwasher", 2007, "Kitchen", "Bosch SHE3AR55UC",
     "SN-BS6621SHEC", "HM-5004",
     1, 12, 650, "F", 2019, 750, 0,
     "Lower spray arm cracked; door seal replaced 2012"),
    ("Toaster Oven", 2012, "Kitchen", "Cuisinart TOB-135",
     "SN-CU4422TOBB", "HM-5005",
     1, 8, 80, "G", 2020, 90, 0,
     "All functions working normally"),

    # ── Laundry (HM-6000s) ───────────────────────────────────────────────────
    ("Washing Machine", 2009, "Laundry Room", "Maytag MVWC360AW",
     "SN-MY5533MVWC", "HM-6001",
     1, 12, 550, "G", 2021, 650, 0,
     "Top-load; no issues; lid switch replaced 2014"),
    ("Clothes Dryer", 2009, "Laundry Room", "Maytag MEDC465HW",
     "SN-MY5534MEDC", "HM-6002",
     1, 12, 500, "G", 2021, 600, 0,
     "Vent duct cleaned annually; heating element good"),

    # ── Garden (HM-7000s) ─────────────────────────────────────────────────────
    ("Gas Lawn Mower", 2006, "Garage / Garden Shed", "Honda HRR216VKA",
     "SN-HO3321HRRM", "HM-7001",
     1, 12, 380, "F", 2018, 420, 0,
     "Carburetor cleaned 2013; self-propel cable fraying"),

    # ── Electronics / Media (HM-8000s) ───────────────────────────────────────
    ("Desktop Computer", 2012, "Home Office", "Dell Inspiron 3847",
     "SN-DL8821INSP", "HM-8001",
     1, 5, 750, "G", 2017, 800, 0,
     "Upgraded RAM to 8 GB 2014; SSD added 2014"),
    ("Flat Panel Television", 2011, "Living Room", "Samsung UN46EH5000",
     "SN-SA6641TVHD", "HM-8002",
     1, 8, 750, "G", 2019, 700, 0,
     "46-inch LED; remote replaced 2013"),
    ("Flat Panel Television", 2009, "Master Bedroom", "LG 32LD450",
     "SN-LG3312BDTV", "HM-8003",
     1, 8, 350, "F", 2017, 320, 0,
     "32-inch LCD; slight backlight bleed in corner"),
    ("Gaming Console", 2013, "Living Room", "Sony PlayStation 3 Slim",
     "SN-SO7711PS3S", "HM-8004",
     1, 8, 250, "G", 2021, 0, 0,
     "Blu-ray drive working; controllers need new analog sticks"),
    ("Home Sound System", 2007, "Living Room", "Yamaha YHT-498",
     "SN-YM2244YHTS", "HM-8005",
     1, 12, 450, "F", 2019, 500, 0,
     "Rear surround speaker has intermittent dropout"),
    ("Laptop Computer", 2010, "Home Office", "HP Pavilion dv6-3200",
     "SN-HP5521PAVDV", "HM-8006",
     1, 5, 600, "P", 2015, 0, 0,
     "Battery no longer holds charge; used as desktop replacement"),

    # ── Furniture (HM-9000s) ─────────────────────────────────────────────────
    ("Sofa / Couch", 2005, "Living Room", "Ashley Darcy Cobblestone",
     "SN-AS1122DCSO", "HM-9001",
     1, 15, 900, "F", 2020, 1000, 0,
     "Seat cushions flattening; fabric fading on arm rests"),
    ("Dining Table and Chair Set", 2003, "Dining Room", "Ethan Allen 122",
     "SN-EA3344DT6C", "HM-9002",
     1, 20, 1400, "F", 2018, 1600, 100,
     "6-chair set; one chair leg repaired 2011; surface scratch on table"),
    ("King Bed Frame", 2008, "Master Bedroom", "IKEA Hemnes King",
     "SN-IK7722HMKB", "HM-9003",
     1, 20, 450, "G", 2028, 500, 0,
     "Solid wood; slats replaced 2013; no structural issues"),
    ("Dresser with Mirror", 2006, "Master Bedroom", "Ashley Signature B376",
     "SN-AS5566DRM3", "HM-9004",
     1, 20, 680, "G", 2026, 750, 50,
     "All drawers slide smoothly; minor scratch on top surface"),
    ("Bookshelf Unit", 2004, "Home Office", "Bush Furniture WC31345",
     "SN-BF9901WSBS", "HM-9005",
     1, 15, 220, "F", 2019, 250, 0,
     "Particle board shelves; one shelf sagging under book load"),

    # ── Outdoor / Sports (HM-9500s) ──────────────────────────────────────────
    ("Fishing Rod Set", 2007, "Garage", "Ugly Stik GX2 Combo",
     "SN-US4411GX2R", "HM-9501",
     2, 15, 90, "G", 2022, 100, 0,
     "Two rod and reel combos; line replaced 2014"),
    ("Camping Tent", 2010, "Garage", "Coleman Sundome 4-Person",
     "SN-CO6622SDTNT", "HM-9502",
     1, 10, 80, "G", 2020, 90, 0,
     "Poles intact; rain fly seams re-sealed 2013"),
    ("Camping Stove", 2008, "Garage", "Coleman Classic 2-Burner",
     "SN-CO3311CSST", "HM-9503",
     1, 15, 50, "G", 2023, 60, 0,
     "Propane stove; burner valves cleaned annually"),
    ("Snow Blower", 2005, "Garage", "Toro Power Clear 518 ZE",
     "SN-TR8821TPSB", "HM-9504",
     1, 12, 420, "F", 2017, 480, 0,
     "Single-stage; auger paddles worn; starter cord fraying"),
    ("Patio Table and Chair Set", 2009, "Backyard Patio", "Hampton Bay Mix & Match",
     "SN-HB2241PTCS", "HM-9505",
     1, 12, 350, "F", 2021, 400, 0,
     "4-chair aluminum set; umbrella faded; light surface rust on table frame"),
]

# ── Write rows ─────────────────────────────────────────────────────────────────
CURRENCY_FMT = '$#,##0'
arial10 = Font(name="Arial", size=10)
center_align = Alignment(horizontal="center")

DATA_START = 7  # first data row

for idx, asset in enumerate(assets, 1):
    (desc, yr_purchase, location, make_model, serial, asset_no,
     qty, dep_life, cost, condition, repl_year, repl_cost_2016, salvage, comments) = asset

    r = DATA_START + idx - 1  # Excel row number

    # Determine condition cell
    # Depreciation formula  =Jn/In
    dep_formula = f"=J{r}/I{r}"
    # Current value formula =MAX(Jn-Ln*(Kn-Cn),0)
    curr_val_formula = f"=MAX(J{r}-L{r}*(K{r}-C{r}),0)"

    row_data = [
        idx,            # A  Entry No.
        desc,           # B  Description
        yr_purchase,    # C  Year of Purchase
        location,       # D  Location
        make_model,     # E  Make and Model
        serial,         # F  Serial No.
        asset_no,       # G  Asset No.
        qty,            # H  Quantity
        dep_life,       # I  Depreciable Life
        cost,           # J  Cost of Asset
        2015,           # K  Year of Update (hardcoded)
        dep_formula,    # L  Depreciation Amount
        curr_val_formula,  # M  Current Value
        condition,      # N  Condition
        repl_year,      # O  Expected Replacement Year
        repl_cost_2016, # P  2016 Replacement Cost
        salvage,        # Q  Expected Salvage
        comments,       # R  Comments
    ]

    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=r, column=col_idx, value=value)
        cell.font = arial10

        col_letter = get_column_letter(col_idx)

        # Currency format: J, L, M, P, Q
        if col_letter in ("J", "L", "M", "P", "Q"):
            cell.number_format = CURRENCY_FMT

        # Center-align N (condition)
        if col_letter == "N":
            cell.alignment = center_align

# ── Freeze header row ─────────────────────────────────────────────────────────
ws.freeze_panes = "A7"

# ── Auto-size columns ─────────────────────────────────────────────────────────
for col in ws.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        try:
            val = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, len(val))
        except Exception:
            pass
    adjusted = min(max_len + 4, 50)
    ws.column_dimensions[col_letter].width = adjusted

# ── Save ──────────────────────────────────────────────────────────────────────
out_path = '/Users/nathandemordaunt/utah-deals/home_asset_register.xlsx'
wb.save(out_path)
print(f"Saved: {out_path}")

# ── Copy to output dir ────────────────────────────────────────────────────────
dest_dir = '/mnt/user-data/outputs'
os.makedirs(dest_dir, exist_ok=True)
dest_path = os.path.join(dest_dir, 'home_asset_register.xlsx')
shutil.copy2(out_path, dest_path)
print(f"Copied to: {dest_path}")
